import os
import re
import json
import tempfile
import logging
from collections import OrderedDict
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel
from typing import Optional, List

from app.services.targets_service import TargetsService
from app.api.routes.auth import get_current_user
from app.database import get_db, SessionLocal
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


router = APIRouter(prefix="/targets", tags=["targets"])


# === Excel 파싱 유틸리티 ===

MONTH_HEADERS = [f"{m}월" for m in range(1, 13)]
SKIP_KEYWORDS = {"소계", "합계", "Total", "total", "TOTAL"}


def _parse_cell_value(value) -> int:
    """셀 값을 정수로 변환 (쉼표, 공백 등 제거)"""
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    try:
        return int(float(cleaned)) if cleaned else 0
    except ValueError:
        return 0


def _should_skip_row(row_values: list) -> bool:
    """소계/합계/Total 행인지 또는 빈 행인지 판별"""
    for val in row_values:
        if val is not None and str(val).strip() in SKIP_KEYWORDS:
            return True
    return False


def _parse_excel_sheet(wb):
    """워크북에서 적절한 시트를 찾아 파싱한 결과를 반환.

    Returns:
        list of dicts, each with keys: manager, category, channel, monthly (list of 12 values)
    """
    from openpyxl.utils import get_column_letter

    parsed_rows = []

    for ws in wb.worksheets:
        # 1) 헤더 행 자동 탐색: "담당자"가 있고, "판매채널" 또는 "채널"이 있는 행
        header_row_idx = None
        col_manager = None
        col_category = None  # 구분
        col_channel = None   # 판매채널
        month_cols = {}      # {0: col_index_for_1월, 1: col_index_for_2월, ...}

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=False), start=1):
            cell_values = [cell.value for cell in row]
            str_values = [str(v).strip() if v is not None else "" for v in cell_values]

            has_manager = "담당자" in str_values
            has_channel = any(v in ("판매채널", "채널") for v in str_values)

            if has_manager and has_channel:
                header_row_idx = row_idx
                for ci, sv in enumerate(str_values):
                    if sv == "담당자":
                        col_manager = ci
                    elif sv == "구분":
                        col_category = ci
                    elif sv in ("판매채널", "채널"):
                        col_channel = ci
                    else:
                        for mi, mh in enumerate(MONTH_HEADERS):
                            if sv == mh:
                                month_cols[mi] = ci
                break

        if header_row_idx is None:
            continue  # 이 시트에는 해당 헤더가 없으므로 다음 시트로

        if col_manager is None or col_channel is None:
            continue

        # 12개월 컬럼이 모두 있는지 확인 (없는 월은 None으로 처리)
        if len(month_cols) == 0:
            continue

        # 2) 데이터 행 읽기
        current_manager = None
        current_category = None

        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=False):
            cell_values = [cell.value for cell in row]

            # 빈 행 스킵: 판매채널 열이 비어있으면 스킵
            channel_val = cell_values[col_channel] if col_channel < len(cell_values) else None
            if channel_val is None or str(channel_val).strip() == "":
                continue

            # 소계/합계/Total 행 스킵
            if _should_skip_row(cell_values):
                continue

            # 담당자: 병합 셀이면 이전 값 유지
            manager_val = cell_values[col_manager] if col_manager < len(cell_values) else None
            if manager_val is not None and str(manager_val).strip():
                current_manager = str(manager_val).strip()

            # 구분: 병합 셀이면 이전 값 유지
            if col_category is not None:
                category_val = cell_values[col_category] if col_category < len(cell_values) else None
                if category_val is not None and str(category_val).strip():
                    current_category = str(category_val).strip()
            else:
                current_category = ""

            channel_str = str(channel_val).strip()

            # 월별 데이터 추출
            monthly = []
            for mi in range(12):
                if mi in month_cols:
                    ci = month_cols[mi]
                    val = cell_values[ci] if ci < len(cell_values) else None
                    monthly.append(_parse_cell_value(val))
                else:
                    monthly.append(0)

            parsed_rows.append({
                "manager": current_manager or "",
                "category": current_category or "",
                "channel": channel_str,
                "monthly": monthly,
            })

        # 첫 번째 유효한 시트를 찾으면 종료
        if parsed_rows:
            break

    return parsed_rows


def _group_by_manager(parsed_rows: list) -> OrderedDict:
    """파싱된 행을 담당자별로 그룹핑.

    Returns:
        OrderedDict: manager_name -> list of row dicts
    """
    groups: OrderedDict = OrderedDict()
    for row in parsed_rows:
        mgr = row["manager"]
        if mgr not in groups:
            groups[mgr] = []
        groups[mgr].append(row)
    return groups


class TargetCreate(BaseModel):
    department: str
    year: int
    title: str
    manager: str
    kpi_type: str  # 매출, 광고선전비, 공헌이익, 판매량
    grid_data: list[list]  # 2D 배열


class TargetUpdate(BaseModel):
    department: Optional[str] = None
    year: Optional[int] = None
    title: Optional[str] = None
    manager: Optional[str] = None
    kpi_type: Optional[str] = None
    grid_data: Optional[list[list]] = None


class SaleCreate(BaseModel):
    year: int
    month: int
    title: str
    manager: str
    kpi_type: str
    grid_data: list[list]


class SaleUpdate(BaseModel):
    year: Optional[int] = None
    month: Optional[int] = None
    title: Optional[str] = None
    manager: Optional[str] = None
    kpi_type: Optional[str] = None
    grid_data: Optional[list[list]] = None


def get_targets_service() -> TargetsService:
    return TargetsService()


# === 목표 데이터 API ===
@router.get("/")
async def get_targets(
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """모든 목표 데이터 조회"""
    if year:
        return service.get_targets_by_year_month(year)
    return service.get_all_targets()


@router.get("/summary")
async def get_target_summary(
    year: int,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """목표 합계 조회"""
    return service.get_target_summary(year, month)


@router.get("/chart/by-manager")
async def get_targets_by_manager(
    year: int,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """책임자별 목표 데이터 조회"""
    return service.get_targets_by_manager(year, month)


@router.get("/chart/by-criteria")
async def get_targets_by_criteria(
    year: int,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """기준별 목표 데이터 조회"""
    return service.get_targets_by_criteria(year, month)


@router.post("/preview-excel")
async def preview_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """엑셀 파일을 파싱하여 미리보기 데이터 반환"""
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다")

    tmp_path = None
    try:
        # 임시 파일로 저장
        suffix = os.path.splitext(file.filename)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name
            content = await file.read()
            tmp.write(content)

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        parsed_rows = _parse_excel_sheet(wb)
        wb.close()

        if not parsed_rows:
            raise HTTPException(status_code=400, detail="엑셀에서 유효한 데이터를 찾을 수 없습니다. 헤더에 '담당자', '판매채널' 열이 필요합니다.")

        groups = _group_by_manager(parsed_rows)

        managers = []
        for mgr_name, rows in groups.items():
            channels = []
            for r in rows:
                channels.append({
                    "category": r["category"],
                    "channel": r["channel"],
                    "monthly": r["monthly"],
                })
            managers.append({
                "manager": mgr_name,
                "channels": channels,
            })

        return {
            "managers": managers,
            "total_channels": len(parsed_rows),
            "total_managers": len(groups),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 파싱 중 오류가 발생했습니다: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    department: str = Form(...),
    year: int = Form(...),
    kpi_type: str = Form(...),
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service),
):
    """엑셀 파일을 파싱하여 담당자별 목표 데이터를 일괄 생성"""
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다")

    tmp_path = None
    try:
        # 임시 파일로 저장
        suffix = os.path.splitext(file.filename)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name
            content = await file.read()
            tmp.write(content)

        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        parsed_rows = _parse_excel_sheet(wb)
        wb.close()

        if not parsed_rows:
            raise HTTPException(status_code=400, detail="엑셀에서 유효한 데이터를 찾을 수 없습니다. 헤더에 '담당자', '판매채널' 열이 필요합니다.")

        groups = _group_by_manager(parsed_rows)

        created_targets = []
        for mgr_name, rows in groups.items():
            # grid_data 구성: 각 행 = ["구분-판매채널", 1월값, 2월값, ..., 12월값]
            grid_data = []
            for r in rows:
                label = f"{r['category']}-{r['channel']}" if r["category"] else r["channel"]
                grid_row = [label] + r["monthly"]
                grid_data.append(grid_row)

            target_data = {
                "department": department,
                "year": year,
                "title": f"{year}년 {mgr_name} 매출 목표",
                "manager": mgr_name,
                "kpi_type": kpi_type,
                "grid_data": grid_data,
            }
            created = service.create_target(target_data)
            created_targets.append(created)

        return created_targets
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 처리 중 오류가 발생했습니다: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.get("/{target_id}")
async def get_target(
    target_id: str,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """특정 목표 데이터 조회"""
    target = service.get_target_by_id(target_id)
    if not target:
        raise HTTPException(status_code=404, detail="목표 데이터를 찾을 수 없습니다")
    return target


@router.post("/")
async def create_target(
    data: TargetCreate,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """새 목표 데이터 생성"""
    return service.create_target(data.model_dump())


@router.put("/{target_id}")
async def update_target(
    target_id: str,
    data: TargetUpdate,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """목표 데이터 수정"""
    target = service.update_target(target_id, data.model_dump(exclude_unset=True))
    if not target:
        raise HTTPException(status_code=404, detail="목표 데이터를 찾을 수 없습니다")
    return target


@router.delete("/{target_id}")
async def delete_target(
    target_id: str,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """목표 데이터 삭제"""
    if not service.delete_target(target_id):
        raise HTTPException(status_code=404, detail="목표 데이터를 찾을 수 없습니다")
    return {"message": "삭제되었습니다"}


# === 매출 현황 API ===
@router.get("/sales/list")
async def get_sales(
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """매출 현황 데이터 조회"""
    if year:
        return service.get_sales_by_year_month(year, month)
    return service.get_all_sales()


@router.get("/sales/summary")
async def get_sales_summary(
    year: int,
    month: int,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """매출 현황 합계 조회"""
    return service.get_sales_summary(year, month)


@router.get("/sales/comparison")
async def get_comparison(
    year: int,
    month: int,
    manager: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """목표 대비 실적, 전월 대비 데이터 조회 (책임자 필터 옵션)"""
    return service.get_comparison_data(year, month, manager)


@router.get("/sales/realtime")
async def get_realtime_indicator(
    year: int,
    month: int,
    manager: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """실시간 지표 조회 (당일 기준 목표 대비 달성률)"""
    return service.get_realtime_indicator(year, month, manager)


@router.get("/sales/daily-chart")
async def get_daily_sales_chart(
    year: int,
    month: int,
    manager: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """일별 매출 데이터 (그래프용)"""
    return service.get_daily_sales_data(year, month, manager)


@router.get("/sales/daily-target")
async def get_daily_target_chart(
    year: int,
    month: int,
    manager: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """일별 목표 데이터 (그래프용) - 월 목표를 일수로 균등 분할"""
    return service.get_daily_target_data(year, month, manager)


@router.get("/sales/chart/by-manager")
async def get_sales_by_manager(
    year: int,
    month: int,
    until_today: bool = True,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """책임자별 매출 데이터 조회 (당일 기준)"""
    sales_data = service.get_sales_by_manager(year, month, until_today)
    target_data = service.get_targets_by_manager_until_day(year, month)
    return {
        "sales": sales_data,
        "targets": target_data,
    }


@router.get("/sales/chart/by-criteria")
async def get_sales_by_criteria(
    year: int,
    month: int,
    until_today: bool = True,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """기준별 매출 데이터 조회 (당일 기준)"""
    sales_data = service.get_sales_by_criteria(year, month, until_today)
    target_data = service.get_targets_by_criteria_until_day(year, month)
    return {
        "sales": sales_data,
        "targets": target_data,
    }


@router.get("/sales/{sale_id}")
async def get_sale(
    sale_id: str,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """특정 매출 현황 데이터 조회"""
    sale = service.get_sale_by_id(sale_id)
    if not sale:
        raise HTTPException(status_code=404, detail="매출 현황 데이터를 찾을 수 없습니다")
    return sale


@router.post("/sales")
async def create_sale(
    data: SaleCreate,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """새 매출 현황 데이터 생성"""
    return service.create_sale(data.model_dump())


@router.put("/sales/{sale_id}")
async def update_sale(
    sale_id: str,
    data: SaleUpdate,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """매출 현황 데이터 수정"""
    sale = service.update_sale(sale_id, data.model_dump(exclude_unset=True))
    if not sale:
        raise HTTPException(status_code=404, detail="매출 현황 데이터를 찾을 수 없습니다")
    return sale


@router.delete("/sales/{sale_id}")
async def delete_sale(
    sale_id: str,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service)
):
    """매출 현황 데이터 삭제"""
    if not service.delete_sale(sale_id):
        raise HTTPException(status_code=404, detail="매출 현황 데이터를 찾을 수 없습니다")
    return {"message": "삭제되었습니다"}


# === 리포트 스케줄 관리 ===

class ReportScheduleCreate(BaseModel):
    name: str = "영업 지표 리포트"
    recipients: str  # comma-separated emails
    schedule_days: str  # comma-separated: 월,화,수
    schedule_time: str = "09:00"
    year: int
    month: Optional[int] = None
    auto_send: bool = True


class ReportScheduleUpdate(BaseModel):
    name: Optional[str] = None
    recipients: Optional[str] = None
    schedule_days: Optional[str] = None
    schedule_time: Optional[str] = None
    year: Optional[int] = None
    month: Optional[int] = None
    auto_send: Optional[bool] = None


@router.get("/report/schedules")
async def list_report_schedules(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """리포트 이메일 스케줄 목록 조회"""
    from app.db_models import TargetReportSchedule
    schedules = db.query(TargetReportSchedule).filter(
        TargetReportSchedule.user_id == current_user.get("sub", "")
    ).order_by(TargetReportSchedule.created_at.desc()).all()
    return [
        {
            "id": s.id,
            "name": s.name,
            "recipients": s.recipients,
            "schedule_days": s.schedule_days,
            "schedule_time": s.schedule_time,
            "year": s.year,
            "month": s.month,
            "auto_send": s.auto_send,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in schedules
    ]


@router.post("/report/schedules")
async def create_report_schedule(
    data: ReportScheduleCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """리포트 이메일 스케줄 생성"""
    from app.db_models import TargetReportSchedule
    schedule = TargetReportSchedule(
        name=data.name,
        recipients=data.recipients,
        schedule_days=data.schedule_days,
        schedule_time=data.schedule_time,
        year=data.year,
        month=data.month,
        auto_send=data.auto_send,
        user_id=current_user.get("sub", ""),
    )
    db.add(schedule)
    db.commit()
    db.refresh(schedule)
    return {
        "id": schedule.id,
        "name": schedule.name,
        "recipients": schedule.recipients,
        "schedule_days": schedule.schedule_days,
        "schedule_time": schedule.schedule_time,
        "year": schedule.year,
        "month": schedule.month,
        "auto_send": schedule.auto_send,
    }


@router.put("/report/schedules/{schedule_id}")
async def update_report_schedule(
    schedule_id: int,
    data: ReportScheduleUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """리포트 이메일 스케줄 수정"""
    from app.db_models import TargetReportSchedule
    schedule = db.query(TargetReportSchedule).filter(
        TargetReportSchedule.id == schedule_id,
        TargetReportSchedule.user_id == current_user.get("sub", ""),
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(schedule, key, value)
    db.commit()
    return {"message": "수정되었습니다"}


@router.delete("/report/schedules/{schedule_id}")
async def delete_report_schedule(
    schedule_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """리포트 이메일 스케줄 삭제"""
    from app.db_models import TargetReportSchedule
    schedule = db.query(TargetReportSchedule).filter(
        TargetReportSchedule.id == schedule_id,
        TargetReportSchedule.user_id == current_user.get("sub", ""),
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="스케줄을 찾을 수 없습니다")
    db.delete(schedule)
    db.commit()
    return {"message": "삭제되었습니다"}


@router.post("/report/schedules/test-now")
async def test_schedule_now(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    service: TargetsService = Depends(get_targets_service),
):
    """등록된 스케줄을 즉시 실행하여 이메일 발송 테스트"""
    from app.db_models import TargetReportSchedule
    from datetime import timedelta, timezone as tz

    KST = tz(timedelta(hours=9))
    now = datetime.now(KST)

    schedules = db.query(TargetReportSchedule).filter(
        TargetReportSchedule.auto_send == True,
    ).all()

    if not schedules:
        return {"error": "활성화된 스케줄이 없습니다", "kst_now": now.strftime("%Y-%m-%d %H:%M:%S")}

    results = []
    for sched in schedules:
        sched_info = {
            "id": sched.id,
            "name": sched.name,
            "recipients": sched.recipients,
            "schedule_days": sched.schedule_days,
            "schedule_time": sched.schedule_time,
            "year": sched.year,
            "month": sched.month,
            "user_id": sched.user_id,
        }

        recipients = [e.strip() for e in (sched.recipients or "").split(",") if e.strip()]
        if not recipients:
            sched_info["status"] = "SKIP: 수신자 없음"
            results.append(sched_info)
            continue

        try:
            from app.config import get_settings
            import resend

            settings = get_settings()
            if not settings.RESEND_API_KEY:
                sched_info["status"] = "SKIP: RESEND_API_KEY 없음"
                results.append(sched_info)
                continue

            resend.api_key = settings.RESEND_API_KEY

            year = sched.year or now.year
            month = sched.month or now.month

            targets = service.get_targets_by_year_month(year, month)
            sales = service.get_sales_by_year_month(year, month)
            by_mgr_target = service.get_targets_by_manager_until_day(year, month)
            by_mgr_sales = service.get_sales_by_manager(year, month, until_today=True)
            html = _build_report_html(year, month, targets, sales, by_mgr_target, by_mgr_sales)

            resend.Emails.send({
                "from": settings.RESEND_FROM_EMAIL,
                "to": recipients,
                "subject": f"[Nuldam] {year}년 {month}월 영업 지표 리포트 (테스트)",
                "html": html,
            })

            sched.last_sent_at = now.replace(tzinfo=None)
            db.commit()
            sched_info["status"] = f"SUCCESS: {recipients} 에게 발송 완료"

        except Exception as e:
            sched_info["status"] = f"ERROR: {str(e)}"

        results.append(sched_info)

    return {
        "kst_now": now.strftime("%Y-%m-%d %H:%M:%S %A"),
        "total_schedules": len(schedules),
        "results": results,
    }


@router.get("/report/schedules/debug")
async def debug_schedules(
    db: Session = Depends(get_db),
):
    """스케줄 디버그 정보 조회"""
    from app.db_models import TargetReportSchedule
    from datetime import timedelta, timezone as tz

    KST = tz(timedelta(hours=9))
    now = datetime.now(KST)
    DAY_MAP = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}
    WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

    schedules = db.query(TargetReportSchedule).all()

    items = []
    for s in schedules:
        days_str = s.schedule_days or ""
        sched_days = [DAY_MAP.get(d.strip(), -1) for d in days_str.split(",") if d.strip()]
        today_match = now.weekday() in sched_days
        time_match = now.strftime("%H:%M") == (s.schedule_time or "09:00")

        items.append({
            "id": s.id,
            "name": s.name,
            "recipients": s.recipients,
            "schedule_days": s.schedule_days,
            "schedule_days_parsed": sched_days,
            "schedule_time": s.schedule_time,
            "auto_send": s.auto_send,
            "user_id": s.user_id,
            "year": s.year,
            "month": s.month,
            "last_sent_at": str(s.last_sent_at) if s.last_sent_at else None,
            "updated_at": str(s.updated_at) if s.updated_at else None,
            "today_day_match": today_match,
            "current_time_match": time_match,
            "would_fire_now": today_match and time_match and s.auto_send,
        })

    return {
        "kst_now": now.strftime("%Y-%m-%d %H:%M:%S"),
        "kst_weekday": WEEKDAY_KR[now.weekday()],
        "total_schedules": len(schedules),
        "schedules": items,
    }


class ReportSendRequest(BaseModel):
    year: int
    month: int
    recipients: str  # comma-separated
    html_content: Optional[str] = None


@router.post("/report/send")
async def send_report_email(
    body: ReportSendRequest,
    current_user: dict = Depends(get_current_user),
    service: TargetsService = Depends(get_targets_service),
):
    """리포트를 이메일로 즉시 발송"""
    try:
        from app.config import get_settings
        import resend

        settings = get_settings()
        if not settings.RESEND_API_KEY:
            raise HTTPException(status_code=400, detail="이메일 설정이 되어있지 않습니다 (RESEND_API_KEY)")

        resend.api_key = settings.RESEND_API_KEY

        # 프론트엔드에서 HTML을 보냈으면 그대로 사용, 아니면 서버에서 생성
        if body.html_content:
            html = body.html_content
        else:
            targets = service.get_targets_by_year_month(body.year, body.month)
            sales = service.get_sales_by_year_month(body.year, body.month)
            by_manager_target = service.get_targets_by_manager_until_day(body.year, body.month)
            by_manager_sales = service.get_sales_by_manager(body.year, body.month, until_today=True)
            html = _build_report_html(body.year, body.month, targets, sales, by_manager_target, by_manager_sales)

        email_list = [e.strip() for e in body.recipients.split(",") if e.strip()]
        if not email_list:
            raise HTTPException(status_code=400, detail="수신자 이메일이 필요합니다")

        resend.Emails.send({
            "from": settings.RESEND_FROM_EMAIL,
            "to": email_list,
            "subject": f"[Nuldam] {body.year}년 {body.month}월 영업 지표 리포트",
            "html": html,
        })
        return {"message": f"{len(email_list)}명에게 리포트를 발송했습니다"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Report email send failed: {e}")
        raise HTTPException(status_code=500, detail=f"이메일 발송 실패: {str(e)}")


def _build_report_html(year, month, targets, sales, by_manager_target, by_manager_sales):
    """리포트 HTML 생성 — 프론트엔드 buildReportHTML()과 동일한 3-page 구조"""
    import calendar
    import re as _re

    # ------------------------------------------------------------------
    # Utility helpers (mirror frontend parseNum / formatKRW / rate colors)
    # ------------------------------------------------------------------
    def _parse_num(val):
        if val is None:
            return 0
        s = _re.sub(r",", "", str(val)).strip()
        try:
            return float(s) if s else 0
        except (ValueError, TypeError):
            return 0

    def _fmt(num):
        """Korean number format with commas (round to int)."""
        return f"{round(num):,}"

    def _rc(rate):
        """Rate color."""
        if rate >= 100:
            return "#10b981"
        if rate >= 80:
            return "#f59e0b"
        return "#ef4444"

    def _rbg(rate):
        """Rate background."""
        if rate >= 100:
            return "#ecfdf5"
        if rate >= 80:
            return "#fffbeb"
        return "#fef2f2"

    def _rbd(rate):
        """Rate border."""
        if rate >= 100:
            return "#a7f3d0"
        if rate >= 80:
            return "#fde68a"
        return "#fecaca"

    # ------------------------------------------------------------------
    # Channel suggestion mapping (same as frontend CHANNEL_SUGGESTIONS)
    # ------------------------------------------------------------------
    CHANNEL_SUGGESTIONS = {
        "스마트스토어": {
            "base": "스마트스토어 키워드 광고 ROI 분석 및 상위노출 키워드 최적화 필요",
            "underperform": "스마트스토어 키워드 광고 ROI가 저조합니다. 긴급히 상위노출 키워드 재설정 및 광고비 재배분이 필요합니다.",
            "overperform": "스마트스토어 성과가 우수합니다. 광고 예산 확대 및 추가 키워드 확장을 통해 성장세를 가속화하세요.",
        },
        "쿠팡": {
            "base": "쿠팡 로켓배송 입점율 확대 및 쿠팡 광고 효율 개선 검토",
            "underperform": "쿠팡 실적이 목표 대비 크게 미달합니다. 로켓배송 입점 확대 및 쿠팡 광고 키워드 전면 재검토가 시급합니다.",
            "overperform": "쿠팡 채널 성과가 매우 우수합니다. 로켓배송 SKU 확대 및 프로모션 강화로 매출 극대화를 추진하세요.",
        },
        "11번가": {
            "base": "11번가 딜 프로모션 참여율 확대 및 가격경쟁력 점검",
            "underperform": "11번가 성과가 부진합니다. 딜 프로모션 적극 참여 및 경쟁사 가격 대비 포지셔닝 재점검이 필요합니다.",
            "overperform": "11번가 실적이 목표를 초과했습니다. 프리미엄 딜 및 브랜드관 입점으로 매출 채널을 확장하세요.",
        },
        "카페24": {
            "base": "자사몰 유입 트래픽 분석 및 리타겟팅 광고 효율화",
            "underperform": "자사몰(카페24) 유입이 저조합니다. 리타겟팅 광고 세팅 점검 및 SNS/검색 유입 채널 다각화가 필요합니다.",
            "overperform": "자사몰 성과가 우수합니다. 자사몰 전용 프로모션 및 멤버십 혜택 강화로 충성 고객 확보에 집중하세요.",
        },
        "올리브영": {
            "base": "올리브영 매대 위치 및 프로모션 기획 강화",
            "underperform": "올리브영 실적이 저조합니다. 매대 위치 재협상 및 올리브영 전용 기획세트 출시를 검토하세요.",
            "overperform": "올리브영 채널이 호조세입니다. 입점 매장 확대 및 시즌 한정 기획세트로 추가 성장을 도모하세요.",
        },
        "카카오": {
            "base": "카카오 선물하기/톡스토어 시즌 프로모션 기획",
            "underperform": "카카오 채널 실적이 미달입니다. 선물하기 상위 노출 전략 및 톡스토어 시즌 프로모션 기획이 시급합니다.",
            "overperform": "카카오 채널이 우수한 성과를 보이고 있습니다. 시즌 프로모션 확대 및 카카오 광고 투자 증대를 권장합니다.",
        },
    }

    def _get_channel_suggestion(channel_name, rate):
        for keyword, suggestions in CHANNEL_SUGGESTIONS.items():
            if keyword in channel_name:
                if rate < 80:
                    return suggestions["underperform"]
                if rate > 120:
                    return suggestions["overperform"]
                return suggestions["base"]
        # Fallback
        if rate < 80:
            return f"{channel_name} 채널의 실적이 목표 대비 저조합니다. 채널별 마케팅 전략 재수립 및 프로모션 강화가 필요합니다."
        if rate > 120:
            return f"{channel_name} 채널이 목표를 크게 초과하고 있습니다. 투자 확대 및 성공 요인 분석을 통해 타 채널에도 적용하세요."
        return f"{channel_name} 채널의 현재 실적을 유지하면서 추가 성장 기회를 모색하세요."

    # ------------------------------------------------------------------
    # effective month & yesterday (mirrors frontend logic)
    # ------------------------------------------------------------------
    effective_month = month
    now_dt = datetime.now()
    days_in_month = calendar.monthrange(year, effective_month)[1]

    if year != now_dt.year or effective_month != now_dt.month:
        yesterday = days_in_month
    else:
        yesterday = max(now_dt.day - 1, 1)

    # ------------------------------------------------------------------
    # 1) Manager target vs actual (from by_manager aggregated data)
    # ------------------------------------------------------------------
    managers_target = by_manager_target.get("by_manager", {})
    managers_sales = by_manager_sales.get("by_manager", {})
    all_managers = sorted(set(list(managers_target.keys()) + list(managers_sales.keys())))

    manager_comparisons = []
    for mgr in all_managers:
        t = managers_target.get(mgr, {})
        s = managers_sales.get(mgr, {})
        target_sales = t.get("매출", 0)
        actual_sales = s.get("매출", 0)
        target_qty = t.get("판매량", 0)
        actual_qty = s.get("판매량", 0)
        target_contrib = t.get("공헌이익", 0)
        actual_contrib = s.get("공헌이익", 0)
        target_adv = t.get("광고선전비", 0)
        actual_adv = s.get("광고선전비", 0) or s.get("마케팅비", 0)

        manager_comparisons.append({
            "manager": mgr,
            "target_sales": target_sales,
            "actual_sales": actual_sales,
            "sales_rate": round(actual_sales / target_sales * 100, 1) if target_sales > 0 else 0,
            "target_qty": target_qty,
            "actual_qty": actual_qty,
            "qty_rate": round(actual_qty / target_qty * 100, 1) if target_qty > 0 else 0,
            "target_contrib": target_contrib,
            "actual_contrib": actual_contrib,
            "contrib_rate": round(actual_contrib / target_contrib * 100, 1) if target_contrib > 0 else 0,
            "target_adv": target_adv,
            "actual_adv": actual_adv,
        })

    # Summary totals
    sum_target_sales = sum(mc["target_sales"] for mc in manager_comparisons)
    sum_actual_sales = sum(mc["actual_sales"] for mc in manager_comparisons)
    sum_target_qty = sum(mc["target_qty"] for mc in manager_comparisons)
    sum_actual_qty = sum(mc["actual_qty"] for mc in manager_comparisons)
    sum_target_contrib = sum(mc["target_contrib"] for mc in manager_comparisons)
    sum_actual_contrib = sum(mc["actual_contrib"] for mc in manager_comparisons)
    sum_sales_rate = round(sum_actual_sales / sum_target_sales * 100, 1) if sum_target_sales > 0 else 0
    sum_qty_rate = round(sum_actual_qty / sum_target_qty * 100, 1) if sum_target_qty > 0 else 0
    sum_contrib_rate = round(sum_actual_contrib / sum_target_contrib * 100, 1) if sum_target_contrib > 0 else 0

    # ------------------------------------------------------------------
    # 2) Channel achievements by manager (from raw targets/sales grid_data)
    # ------------------------------------------------------------------
    channel_achievements = []
    manager_set = sorted(set(
        [t.get("manager", "") for t in targets] +
        [s.get("manager", "") for s in sales]
    ))

    for mgr in manager_set:
        if not mgr:
            continue
        # Target grid_data for 매출 kpi
        mgr_targets = [t for t in targets if t.get("manager") == mgr and t.get("kpi_type") == "매출"]
        mgr_sales = [s for s in sales if s.get("manager") == mgr and s.get("kpi_type") == "매출"]

        channel_target_map = {}
        for tgt in mgr_targets:
            grid = tgt.get("grid_data", [])
            for row in grid:
                if not row or len(row) == 0:
                    continue
                ch_label = row[0]
                if not ch_label:
                    continue
                # Target grid: [channel, m1, m2, ..., m12]; index = effective_month (1-based)
                month_idx = effective_month
                val = _parse_num(row[month_idx]) if len(row) > month_idx else 0
                channel_target_map[ch_label] = channel_target_map.get(ch_label, 0) + val

        channel_actual_map = {}
        for sl in mgr_sales:
            grid = sl.get("grid_data", [])
            for row in grid:
                if not row or len(row) == 0:
                    continue
                ch_label = row[0]
                if not ch_label:
                    continue
                # Sales grid: [channel, day1, day2, ..., day31]; sum all days
                total = sum(_parse_num(row[i]) for i in range(1, len(row)))
                channel_actual_map[ch_label] = channel_actual_map.get(ch_label, 0) + total

        all_channels = sorted(set(list(channel_target_map.keys()) + list(channel_actual_map.keys())))
        for ch in all_channels:
            t_val = channel_target_map.get(ch, 0)
            a_val = channel_actual_map.get(ch, 0)
            rate = round(a_val / t_val * 100, 1) if t_val > 0 else (999 if a_val > 0 else 0)
            channel_achievements.append({
                "manager": mgr,
                "channel": ch,
                "target": t_val,
                "actual": a_val,
                "rate": rate,
                "achieved": rate >= 100,
            })

    # ------------------------------------------------------------------
    # 3) Manager entry status (from raw sales grid_data)
    # ------------------------------------------------------------------
    manager_sales_map = {}
    for sl in sales:
        mgr = sl.get("manager", "")
        if mgr not in manager_sales_map:
            manager_sales_map[mgr] = []
        manager_sales_map[mgr].append(sl)
    # Also include managers from targets who might not have sales yet
    for tgt in targets:
        mgr = tgt.get("manager", "")
        if mgr and mgr not in manager_sales_map:
            manager_sales_map[mgr] = []

    entry_statuses = []
    for mgr in sorted(manager_sales_map.keys()):
        if not mgr:
            continue
        max_day = 0
        for sl in manager_sales_map[mgr]:
            grid = sl.get("grid_data", [])
            for row in grid:
                if not row or len(row) <= 1:
                    continue
                for i in range(len(row) - 1, 0, -1):
                    if _parse_num(row[i]) != 0:
                        if i > max_day:
                            max_day = i
                        break

        is_normal = max_day >= yesterday
        missing_days = 0 if max_day >= yesterday else yesterday - max_day
        entry_statuses.append({
            "manager": mgr,
            "last_entry_day": max_day,
            "required_day": yesterday,
            "is_normal": is_normal,
            "missing_days": missing_days,
        })

    # Flagged managers
    missing_entry_managers = [s for s in entry_statuses if not s["is_normal"]]

    # ------------------------------------------------------------------
    # 4) Channel suggestions (aggregate across managers)
    # ------------------------------------------------------------------
    ch_agg = {}
    for ca in channel_achievements:
        ch = ca["channel"]
        if ch not in ch_agg:
            ch_agg[ch] = {"total_target": 0, "total_actual": 0}
        ch_agg[ch]["total_target"] += ca["target"]
        ch_agg[ch]["total_actual"] += ca["actual"]

    channel_suggestions = []
    for ch, vals in ch_agg.items():
        t_total = vals["total_target"]
        a_total = vals["total_actual"]
        rate = round(a_total / t_total * 100, 1) if t_total > 0 else (999 if a_total > 0 else 0)
        channel_suggestions.append({
            "channel": ch,
            "rate": rate,
            "suggestion": _get_channel_suggestion(ch, rate),
        })
    channel_suggestions.sort(key=lambda x: x["rate"])

    # ==================================================================
    # BUILD HTML (mirrors frontend buildReportHTML exactly)
    # ==================================================================

    # --- Manager comparison rows ---
    mgr_rows = ""
    for mc in manager_comparisons:
        mgr_rows += (
            f'<tr>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;font-weight:600;color:#1e293b">{mc["manager"]}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["target_sales"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["actual_sales"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;font-weight:700;color:{_rc(mc["sales_rate"])}">{mc["sales_rate"]:.1f}%</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["target_qty"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["actual_qty"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;font-weight:700;color:{_rc(mc["qty_rate"])}">{mc["qty_rate"]:.1f}%</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["target_contrib"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(mc["actual_contrib"])}</td>'
            f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;font-weight:700;color:{_rc(mc["contrib_rate"])}">{mc["contrib_rate"]:.1f}%</td>'
            f'</tr>'
        )

    sum_row = (
        f'<tr style="background:#f5f3ff;font-weight:700;border-top:2px solid #94a3b8">'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;color:#1e293b">합계</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_target_sales)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_actual_sales)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:{_rc(sum_sales_rate)}">{sum_sales_rate:.1f}%</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_target_qty)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_actual_qty)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:{_rc(sum_qty_rate)}">{sum_qty_rate:.1f}%</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_target_contrib)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right">{_fmt(sum_actual_contrib)}</td>'
        f'<td style="padding:10px 12px;border:1px solid #e2e8f0;text-align:right;color:{_rc(sum_contrib_rate)}">{sum_contrib_rate:.1f}%</td>'
        f'</tr>'
    )

    # --- CSS bar chart ---
    bars = ""
    for mc in manager_comparisons:
        w = min(mc["sales_rate"], 150)
        bars += (
            f'<div style="display:flex;align-items:center;margin-bottom:8px">'
            f'<div style="width:80px;font-size:13px;font-weight:600;color:#334155;flex-shrink:0">{mc["manager"]}</div>'
            f'<div style="flex:1;display:flex;align-items:center;gap:4px">'
            f'<div style="height:20px;width:{w}%;background:{_rc(mc["sales_rate"])};border-radius:4px;min-width:2px"></div>'
            f'<span style="font-size:12px;font-weight:600;color:{_rc(mc["sales_rate"])}">{mc["sales_rate"]:.1f}%</span>'
            f'</div>'
            f'</div>'
        )

    # --- Channel by manager ---
    grouped_ch = {}
    for ca in channel_achievements:
        mgr = ca["manager"]
        if mgr not in grouped_ch:
            grouped_ch[mgr] = []
        grouped_ch[mgr].append(ca)

    ch_sections = ""
    for mgr in sorted(grouped_ch.keys()):
        channels = grouped_ch[mgr]
        ch_rows = ""
        for ch in sorted(channels, key=lambda x: x["rate"]):
            st_bg = "#ecfdf5" if ch["rate"] >= 100 else "#fef2f2"
            st_c = "#059669" if ch["rate"] >= 100 else "#dc2626"
            st_t = "달성" if ch["rate"] >= 100 else "미달성"
            rate_display = "-" if ch["rate"] > 900 else f'{ch["rate"]:.1f}%'
            ch_rows += (
                f'<tr>'
                f'<td style="padding:8px 12px;border:1px solid #e2e8f0;color:#334155">{ch["channel"]}</td>'
                f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(ch["target"])}</td>'
                f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:right;color:#475569">{_fmt(ch["actual"])}</td>'
                f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:right;font-weight:700;color:{_rc(ch["rate"])}">{rate_display}</td>'
                f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:center">'
                f'<span style="display:inline-block;padding:2px 10px;border-radius:999px;font-size:12px;font-weight:600;background:{st_bg};color:{st_c}">{st_t}</span>'
                f'</td>'
                f'</tr>'
            )
        ch_sections += (
            f'<div style="margin-bottom:16px;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">'
            f'<div style="padding:10px 14px;background:#eff6ff;border-bottom:1px solid #e2e8f0">'
            f'<strong style="color:#334155;font-size:14px">{mgr}</strong>'
            f'</div>'
            f'<table style="width:100%;border-collapse:collapse;font-size:13px">'
            f'<thead><tr style="background:#f8fafc">'
            f'<th style="text-align:left;padding:8px 12px;border:1px solid #e2e8f0;color:#64748b;font-weight:600">채널</th>'
            f'<th style="text-align:right;padding:8px 12px;border:1px solid #e2e8f0;color:#64748b;font-weight:600">목표</th>'
            f'<th style="text-align:right;padding:8px 12px;border:1px solid #e2e8f0;color:#64748b;font-weight:600">실적</th>'
            f'<th style="text-align:right;padding:8px 12px;border:1px solid #e2e8f0;color:#64748b;font-weight:600">달성률</th>'
            f'<th style="text-align:center;padding:8px 12px;border:1px solid #e2e8f0;color:#64748b;font-weight:600">상태</th>'
            f'</tr></thead>'
            f'<tbody>{ch_rows}</tbody>'
            f'</table>'
            f'</div>'
        )

    if not ch_sections:
        ch_sections = '<div style="padding:20px;text-align:center;color:#94a3b8;background:#f8fafc;border-radius:8px">데이터 없음</div>'

    # --- Entry status table ---
    entry_rows = ""
    for st in entry_statuses:
        st_bg = "#ecfdf5" if st["is_normal"] else "#fef2f2"
        st_c = "#059669" if st["is_normal"] else "#dc2626"
        st_t = "정상" if st["is_normal"] else "미입력"
        last_day_text = f'{st["last_entry_day"]}일' if st["last_entry_day"] > 0 else "입력 없음"
        missing_text = f'{st["missing_days"]}일' if st["missing_days"] > 0 else "-"
        missing_color = "#ef4444" if st["missing_days"] > 0 else "#94a3b8"
        entry_rows += (
            f'<tr>'
            f'<td style="padding:8px 12px;border:1px solid #e2e8f0;font-weight:600;color:#1e293b">{st["manager"]}</td>'
            f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:center;color:#475569">{last_day_text}</td>'
            f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:center;color:#475569">{st["required_day"]}일</td>'
            f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:center;font-weight:700;color:{missing_color}">{missing_text}</td>'
            f'<td style="padding:8px 12px;border:1px solid #e2e8f0;text-align:center">'
            f'<span style="display:inline-block;padding:2px 10px;border-radius:999px;font-size:12px;font-weight:600;background:{st_bg};color:{st_c}">{st_t}</span>'
            f'</td>'
            f'</tr>'
        )

    # --- Missing entry warning cards ---
    warn_cards = ""
    for st in missing_entry_managers:
        last_info = f'{effective_month}월 {st["last_entry_day"]}일' if st["last_entry_day"] > 0 else "입력 없음"
        warn_cards += (
            f'<div style="display:inline-block;width:calc(33% - 12px);min-width:200px;margin:4px;padding:14px;border:1px solid #fecaca;border-radius:8px;background:#fff;vertical-align:top">'
            f'<div style="display:flex;align-items:center;gap:10px">'
            f'<div style="width:36px;height:36px;border-radius:50%;background:#fee2e2;display:flex;align-items:center;justify-content:center;flex-shrink:0">'
            f'<span style="font-weight:700;color:#dc2626;font-size:14px">{st["manager"][:1]}</span>'
            f'</div>'
            f'<div>'
            f'<div style="font-weight:600;color:#1e293b;font-size:14px">{st["manager"]}</div>'
            f'<div style="font-size:12px;color:#64748b;margin-top:2px">마지막 입력일: {last_info}</div>'
            f'<div style="font-size:12px;color:#dc2626;font-weight:600;margin-top:2px">미입력 일수: {st["missing_days"]}일</div>'
            f'</div>'
            f'</div>'
            f'</div>'
        )

    # --- Channel suggestions ---
    sug_cards = ""
    for cs in channel_suggestions:
        rate_text = "목표 미설정" if cs["rate"] > 900 else f'달성률 {cs["rate"]:.1f}%'
        sug_cards += (
            f'<div style="padding:14px;border:1px solid {_rbd(cs["rate"])};border-radius:8px;background:{_rbg(cs["rate"])};margin-bottom:10px">'
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">'
            f'<span style="font-weight:700;color:{_rc(cs["rate"])};font-size:14px">{cs["channel"]}</span>'
            f'<span style="font-size:11px;font-weight:700;padding:2px 8px;border-radius:999px;background:{_rbd(cs["rate"])};color:{_rc(cs["rate"])}">{rate_text}</span>'
            f'</div>'
            f'<p style="font-size:13px;line-height:1.5;color:#334155;margin:0">{cs["suggestion"]}</p>'
            f'</div>'
        )

    if not sug_cards:
        sug_cards = '<div style="padding:20px;text-align:center;color:#94a3b8;background:#f8fafc;border-radius:8px">데이터 없음</div>'

    now_str = now_dt.strftime("%Y. %m. %d. %H:%M:%S")
    th_style = 'text-align:right;padding:10px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700'

    # Section number for channel suggestions (shifts if missing-entry section exists)
    missing_section_num = "4" if missing_entry_managers else ""
    sug_section_num = "5" if missing_entry_managers else "4"

    # --- Missing entry section HTML ---
    missing_section = ""
    if missing_entry_managers:
        missing_section = (
            f'<div style="margin-bottom:24px">'
            f'<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;overflow:hidden">'
            f'<div style="padding:14px 16px;border-bottom:1px solid #fecaca;display:flex;align-items:center;gap:8px">'
            f'<span style="font-size:18px;color:#dc2626">&#9888;</span>'
            f'<h2 style="font-size:16px;font-weight:700;color:#991b1b;margin:0">4. 미입력자 경고</h2>'
            f'<span style="padding:2px 10px;background:#fee2e2;color:#dc2626;border-radius:999px;font-size:12px;font-weight:700">{len(missing_entry_managers)}명</span>'
            f'</div>'
            f'<div style="padding:14px">{warn_cards}</div>'
            f'</div>'
            f'</div>'
        )

    # ==================================================================
    # Assemble final HTML
    # ==================================================================
    return (
        f'<div style="font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Roboto,Arial,sans-serif;color:#1e293b;max-width:100%;padding:24px;font-size:13px;line-height:1.6">'

        # PAGE 1
        f'<div style="page-break-after:always">'
        f'<div style="text-align:center;padding-bottom:20px;border-bottom:3px solid #6366f1;margin-bottom:24px">'
        f'<h1 style="font-size:22px;font-weight:800;color:#1e1b4b;margin:0 0 6px 0">{year}년 {effective_month}월 목표 달성 현황 리포트</h1>'
        f'<p style="font-size:13px;color:#64748b;margin:0">기준일: {effective_month}월 {yesterday}일까지 | 생성일시: {now_str}</p>'
        f'</div>'
        f'<div style="margin-bottom:20px">'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:14px">'
        f'<div style="width:4px;height:20px;border-radius:2px;background:#7c3aed"></div>'
        f'<h2 style="font-size:16px;font-weight:700;color:#1e293b;margin:0">1. 담당자별 목표 vs 실적 현황</h2>'
        f'</div>'
        f'<table style="width:100%;border-collapse:collapse;font-size:12px">'
        f'<thead><tr style="background:#f5f3ff">'
        f'<th style="text-align:left;padding:10px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">담당자</th>'
        f'<th style="{th_style}">목표매출</th><th style="{th_style}">실적매출</th><th style="{th_style}">매출달성률</th>'
        f'<th style="{th_style}">목표판매량</th><th style="{th_style}">실적판매량</th><th style="{th_style}">판매량달성률</th>'
        f'<th style="{th_style}">목표공헌이익</th><th style="{th_style}">실적공헌이익</th><th style="{th_style}">공헌이익달성률</th>'
        f'</tr></thead>'
        f'<tbody>{mgr_rows}{sum_row}</tbody>'
        f'</table>'
        f'</div>'
        f'<div style="margin-top:20px;padding:16px;background:#f8fafc;border-radius:8px;border:1px solid #e2e8f0">'
        f'<h3 style="font-size:14px;font-weight:700;color:#334155;margin:0 0 12px 0">매출 달성률 차트</h3>'
        f'{bars}'
        f'<div style="display:flex;gap:16px;margin-top:10px;font-size:11px;color:#64748b">'
        f'<span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#10b981;margin-right:4px"></span>100% 이상</span>'
        f'<span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#f59e0b;margin-right:4px"></span>80~99%</span>'
        f'<span><span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:#ef4444;margin-right:4px"></span>80% 미만</span>'
        f'</div>'
        f'</div>'
        f'</div>'

        # PAGE 2
        f'<div style="page-break-after:always">'
        f'<div style="margin-bottom:24px">'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:14px">'
        f'<div style="width:4px;height:20px;border-radius:2px;background:#3b82f6"></div>'
        f'<h2 style="font-size:16px;font-weight:700;color:#1e293b;margin:0">2. 채널별 달성/미달성 현황</h2>'
        f'</div>'
        f'{ch_sections}'
        f'</div>'
        f'<div>'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:14px">'
        f'<div style="width:4px;height:20px;border-radius:2px;background:#f59e0b"></div>'
        f'<h2 style="font-size:16px;font-weight:700;color:#1e293b;margin:0">3. 담당자별 데이터 입력 현황</h2>'
        f'</div>'
        f'<p style="font-size:12px;color:#64748b;margin:0 0 10px 0">기준일: {year}년 {effective_month}월 {yesterday}일까지 입력 필요</p>'
        f'<table style="width:100%;border-collapse:collapse;font-size:13px">'
        f'<thead><tr style="background:#fffbeb">'
        f'<th style="text-align:left;padding:8px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">담당자</th>'
        f'<th style="text-align:center;padding:8px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">마지막 입력일</th>'
        f'<th style="text-align:center;padding:8px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">요구 입력일</th>'
        f'<th style="text-align:center;padding:8px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">미입력 일수</th>'
        f'<th style="text-align:center;padding:8px 12px;border:1px solid #e2e8f0;color:#475569;font-weight:700">상태</th>'
        f'</tr></thead>'
        f'<tbody>{entry_rows}</tbody>'
        f'</table>'
        f'</div>'
        f'</div>'

        # PAGE 3
        f'<div>'
        f'{missing_section}'
        f'<div>'
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:14px">'
        f'<div style="width:4px;height:20px;border-radius:2px;background:#10b981"></div>'
        f'<h2 style="font-size:16px;font-weight:700;color:#1e293b;margin:0">{sug_section_num}. 채널별 개선 아이디어</h2>'
        f'</div>'
        f'{sug_cards}'
        f'</div>'
        f'<div style="margin-top:30px;padding-top:16px;border-top:1px solid #e2e8f0;text-align:center">'
        f'<p style="font-size:11px;color:#94a3b8;margin:0">이 리포트는 Nuldam Analytics에서 자동 생성되었습니다.</p>'
        f'</div>'
        f'</div>'

        f'</div>'
    )
