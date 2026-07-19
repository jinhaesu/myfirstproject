"""CJ 온스타일 파서. 세 가지 포맷 지원.

1) '인기상품분석현황' 집계 포맷:
  row0: 메타 ("인기상품분석현황")
  row1: 헤더 (순위, 상품코드, 상품명, 총주문수량, 총주문금액)
  row2+: 데이터. 일자 컬럼 없음 → 업로드 일자로 임시 적재.

2) B2B 납품형: 상품명 + 납품예정 금액/납품금액.

3) '배송통합조회' 배송 단위 포맷 (74컬럼):
  No·배송상태·배송지시일·주문번호·상품명·옵션명·수량·결제일·
  공급가(협력사 공급가, VAT별도)·판매가(소비자 단가)·결제가(실결제)·취소예정 …

  기준변경요청서(임현정 MD, 2026-07-14) — 아래가 최종 규칙(구 공급가/판매가×수량
  기준(2026-07-06 임시승인분)은 폐기):
    - 기준일 = AM열[결제일] (출고일 아님).
    - 취소/환불/반품 = AF열[주문상태] 또는 B열[배송상태] 텍스트에 '취소'/'환불'/'반품'
      포함 여부만으로 판별(취소예정 플래그·배송취소일은 더 이상 사용 안 함).
      버리지 않고 is_cancelled로 표시 — 매출·낱개엔 미반영, 건수만 보존.
    - 매출(net) = AW열[결제가] (VAT포함 소비자 실결제액, 수량 반영된 라인 총액).
      ingest 시점에 CJ온스타일이 VAT_INCLUDED_CHANNELS에 속해 자동 ÷1.1 → 여기선
      raw 그대로 넘김(이중 나눗셈 금지).
    - 낱개수량(unit_per_set) = O열[상품명]+P열[옵션명] 결합 텍스트에 정규식
      우선순위(_cj_unit_per_set) 적용 × Q열[수량](시스템이 raw_qty×unit_per_set로 환산).
    6월 샘플 검증: 순매출 3,126,763원 / 환산낱개 1,726개 = 담당자 정답과 정확 일치.

openpyxl CellStyle 버그 우회 → python-calamine 사용.
"""
from __future__ import annotations
import re
from datetime import date
from typing import Iterable, Optional

import pandas as pd

from app.services.csa_service import ParsedLine
from app.services.csa_parsers import register
from app.services.csa_parsers._common import to_date, to_float, to_str


def _read(path: str) -> pd.DataFrame:
    """header=None으로 읽고 상품명/총주문금액이 있는 행을 헤더로 사용."""
    for engine in ("calamine", "openpyxl"):
        try:
            df = pd.read_excel(path, engine=engine, header=None)
            return df
        except Exception:
            continue
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows) if rows else pd.DataFrame()


_VAT = 1.1  # 결제가 컬럼이 없는 구버전 파일에서 공급가 폴백 시 상쇄용
_CANCEL_TOKENS = ("취소", "반품", "환불")

# CJ온스타일 낱개수량(입수) 정규식 우선순위 (기준변경요청서, 임현정 MD, 2026-07-14).
# 상품명+옵션명 결합 텍스트에 순서대로 적용, 첫 매치 채택.
_TOTAL_RE = re.compile(r"총\s*(\d+)\s*(?:개입|구|봉|병|개)")           # 1순위: '총 N구/개/봉/병/개입'
_PLUS_RE = re.compile(r"(\d+)\s*\+\s*(\d+)\s*(?:구|병|개|봉|개입)?")   # 2순위: 'N+N' 합산
_JONG_SET_RE = re.compile(r"(\d+)\s*종\s*(\d+)\s*(?:세트|셋트|box)", re.I)  # 3순위: 'N종 M세트/BOX' → N×M
_UNIT_RE = re.compile(r"(\d+)\s*(?:개입|구|봉|병|개)")                 # 4순위: 'N구/개/봉/병/개입' 최댓값
_JONG_RE = re.compile(r"(\d+)\s*종")                                   # 5순위: 'N종'


def _cj_unit_per_set(prod: Optional[str], opt: Optional[str]) -> float:
    """상품명[O]+옵션명[P] 결합 텍스트 → 세트 입수(N). 매치 없으면 단품 기본값 1(6순위)."""
    t = f"{prod or ''} {opt or ''}"
    m = _TOTAL_RE.search(t)
    if m:
        return float(m.group(1))
    m = _PLUS_RE.search(t)
    if m:
        return float(int(m.group(1)) + int(m.group(2)))
    m = _JONG_SET_RE.search(t)
    if m:
        return float(int(m.group(1)) * int(m.group(2)))
    nums = [int(x) for x in _UNIT_RE.findall(t)]
    nums = [n for n in nums if 1 <= n <= 200]
    if nums:
        return float(max(nums))
    m = _JONG_RE.search(t)
    if m:
        return float(m.group(1))
    return 1.0


def _parse_delivery_report(df: pd.DataFrame, header_row: int) -> Iterable[ParsedLine]:
    """'배송통합조회' 배송 단위 포맷. 규칙은 모듈 docstring 참조."""
    header = [str(x or "").strip() for x in df.iloc[header_row].tolist()]
    for idx in range(header_row + 1, len(df)):
        d = {h: v for h, v in zip(header, df.iloc[idx].tolist())}
        prod = to_str(d.get("상품명"))
        if not prod or prod == "상품명":
            continue
        # 취소/환불/반품 = 주문상태[AF] 또는 배송상태[B] 텍스트 포함 여부만으로 판별.
        # 버리지 않고 is_cancelled로 표시 — 매출·낱개엔 미반영, 건수만 보존(기존 파서 컨벤션).
        state = (to_str(d.get("주문상태")) or "") + (to_str(d.get("배송상태")) or "")
        is_cancel = any(t in state for t in _CANCEL_TOKENS)

        qty = to_float(d.get("수량") or 0)
        supply = to_float(d.get("공급가") or 0)   # 협력사 공급가 (VAT별도) — 구버전 폴백용
        price = to_float(d.get("판매가") or 0)    # 소비자 판매 단가 (VAT포함) — 구버전 폴백용
        paid = to_float(d.get("결제가") or 0)     # 실결제액 (VAT포함, 라인 총액)
        if qty == 0 and supply == 0 and price == 0 and paid == 0:
            continue
        # 순매출 = 결제가(AW) 그대로(라인 총액, VAT포함). ingest에서 CJ온스타일이
        # VAT_INCLUDED_CHANNELS라 자동 ÷1.1 → 여기선 나누지 않는다(이중 나눗셈 금지).
        # 결제가가 없는 구버전 파일은 판매가×수량 → 공급가×1.1 순으로 폴백.
        net = paid or ((price * qty) if price and qty else supply * _VAT)
        opt = to_str(d.get("옵션명"))
        sale_d = (
            to_date(d.get("결제일")) or to_date(d.get("배송지시일"))
            or to_date(d.get("출고일")) or date.today()
        )
        yield ParsedLine(
            sale_date=sale_d,
            order_no=to_str(d.get("주문번호")),
            line_no=f"{to_str(d.get('상품코드')) or ''}-{idx}",
            raw_product_name=prod,
            raw_option_name=opt,
            raw_qty=qty,
            gross_amount=0 if is_cancel else net,
            net_amount=0 if is_cancel else net,
            refund_amount=net if is_cancel else 0,
            is_cancelled=is_cancel,
            unit_per_set=_cj_unit_per_set(prod, opt),
        )


@register("CJ온스타일")
@register("CJ 온스타일")
@register("CJ ON STYLE")
def parse(path: str) -> Iterable[ParsedLine]:
    df = _read(path)
    if df.empty:
        return

    # 헤더 행 찾기 — 집계형(총주문금액) / B2B 납품형(납품예정 금액) /
    # 배송통합조회형(공급가·배송상태) 모두 대응
    header_row = None
    is_delivery = False
    for i in range(min(8, len(df))):
        vals = [str(x or "").strip() for x in df.iloc[i].tolist()]
        has_prod = "상품명" in vals
        if has_prod and "공급가" in vals and "배송상태" in vals:
            header_row = i
            is_delivery = True
            break
        has_amount = any(
            kw in v
            for v in vals
            for kw in ("총주문금액", "주문금액", "총금액", "납품예정 금액", "납품금액")
        )
        if has_prod and has_amount:
            header_row = i
            break
    if header_row is None:
        return
    if is_delivery:
        yield from _parse_delivery_report(df, header_row)
        return

    header = [str(x or "").strip() for x in df.iloc[header_row].tolist()]
    today = date.today()
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx].tolist()
        d = {h: v for h, v in zip(header, row)}
        prod = to_str(d.get("상품명"))
        if not prod or prod == "상품명":
            continue
        qty = to_float(d.get("총주문수량") or d.get("주문수량") or d.get("상품 수량") or d.get("수량") or 0)
        gross = to_float(
            d.get("총주문금액") or d.get("주문금액") or d.get("총금액")
            or d.get("납품예정 금액") or d.get("납품금액") or 0
        )
        if qty == 0 and gross == 0:
            continue
        # 날짜: 집계형은 업로드 일자, B2B형은 납품일자 우선
        from app.services.csa_parsers._common import to_date
        sale_d = (
            to_date(d.get("납품일자") or d.get("발주일자") or d.get("납품예정일자"))
            or today
        )
        yield ParsedLine(
            sale_date=sale_d,
            line_no=to_str(d.get("상품코드") or d.get("납품예정번호")),
            raw_product_name=prod,
            raw_qty=qty,
            gross_amount=gross,
            net_amount=gross,
        )
