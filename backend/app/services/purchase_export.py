"""구매 관리 엑셀(.xlsx) 내보내기.

각 메뉴(실적조회·단가추이·실적대시보드·월별매트릭스·매입채무·BOM매핑)의 주요 내용을
openpyxl 워크북으로 만든다. kind='all'이면 전 메뉴를 시트로 묶어 한 파일로 반환.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.services import purchase_service as pur

_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
_SUM_FILL = PatternFill("solid", fgColor="E5E7EB")
_SUM_FONT = Font(bold=True, size=10)
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0"
_NUM1 = "#,##0.0"


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list],
           money_cols: Optional[set] = None, num1_cols: Optional[set] = None,
           sum_row: Optional[list] = None, first: bool = False):
    """헤더+본문(+합계행)을 가진 시트를 만든다. money_cols/num1_cols는 0-based 열 인덱스."""
    ws = wb.active if first else wb.create_sheet()
    ws.title = title[:31]
    money_cols = money_cols or set()
    num1_cols = num1_cols or set()

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _HEAD_FILL; c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    widths = [max(len(str(h)) + 2, 8) for h in headers]
    r_i = 2
    for row in rows:
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=r_i, column=ci, value=val)
            c.border = _BORDER
            if (ci - 1) in money_cols:
                c.number_format = _MONEY
                c.alignment = Alignment(horizontal="right")
            elif (ci - 1) in num1_cols:
                c.number_format = _NUM1
                c.alignment = Alignment(horizontal="right")
            L = len(str(val)) if val is not None else 0
            if L + 2 > widths[ci - 1]:
                widths[ci - 1] = min(L + 2, 48)
        r_i += 1
    if sum_row is not None:
        for ci, val in enumerate(sum_row, 1):
            c = ws.cell(row=r_i, column=ci, value=val)
            c.fill = _SUM_FILL; c.font = _SUM_FONT; c.border = _BORDER
            if (ci - 1) in money_cols:
                c.number_format = _MONEY
                c.alignment = Alignment(horizontal="right")
            elif (ci - 1) in num1_cols:
                c.number_format = _NUM1
                c.alignment = Alignment(horizontal="right")
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    return ws


# ── kind별 시트 빌더 ──

def _sheet_records(wb, db, params, first=False):
    d = pur.records_list(
        db, start=params.get("start"), end=params.get("end"),
        vendor=params.get("vendor"), mclass=params.get("mclass"),
        item_code=params.get("item_code"), q=params.get("q"),
        team=params.get("team"), limit=100000, offset=0)
    headers = ["일자", "No", "담당팀", "창고", "거래처", "구분", "담당자", "품목코드",
               "품목명", "규격", "단위", "수량", "kg", "단가", "kg당단가",
               "공급가", "부가세", "합계", "정산", "정산일", "비고"]
    rows = []
    for r in d["rows"]:
        rows.append([
            r.get("pdate"), r.get("seq"), r.get("team"), r.get("warehouse"),
            r.get("vendor"), r.get("mclass"), r.get("staff"), r.get("item_code"),
            r.get("item_name_short") or r.get("item_name"), r.get("spec"), r.get("unit"),
            r.get("qty"), r.get("kg"), r.get("unit_price"), r.get("price_per_kg"),
            r.get("supply"), r.get("vat"), r.get("total"),
            "완료" if r.get("paid") else "미지급", r.get("paid_date"), r.get("note"),
        ])
    sum_row = ["합계", "", "", "", "", "", "", "", f"{d['total']}건", "", "",
               d.get("qty_total"), d.get("kg_total"), "", "",
               d.get("supply_total"), d.get("vat_total"), d.get("total_total"), "", "", ""]
    _sheet(wb, "실적조회", headers, rows,
           money_cols={13, 14, 15, 16, 17}, num1_cols={11, 12},
           sum_row=sum_row, first=first)


def _sheet_price(wb, db, params, first=False):
    d = pur.price_tracker(
        db, params["start"], params["end"], mclass=params.get("mclass"),
        vendor=params.get("vendor"), q=params.get("q"),
        min_lines=params.get("min_lines", 1), sort=params.get("sort", "abs_change"),
        team=params.get("team"))
    headers = ["품목코드", "품목명", "규격", "구분", "단위", "거래처수", "매입횟수",
               "최초일", "최초단가", "최근일", "최근단가", "최저", "최고", "평균",
               "변동", "변동률%", "스프레드%", "총수량", "총kg", "총매입액", "단가효과"]
    rows = []
    for x in d["items"]:
        rows.append([
            x.get("item_code"), x.get("item_name_short") or x.get("item_name"), x.get("spec"),
            x.get("mclass"), x.get("unit"), x.get("vendor_count"), x.get("buy_count"),
            x.get("first_date"), x.get("first_price"), x.get("last_date"), x.get("last_price"),
            x.get("min_price"), x.get("max_price"), x.get("avg_price"),
            x.get("change"), x.get("change_pct"), x.get("spread_pct"),
            x.get("total_qty"), x.get("total_kg"), x.get("total_supply"), x.get("cost_impact"),
        ])
    _sheet(wb, "단가추이", headers, rows,
           money_cols={8, 10, 11, 12, 13, 14, 19, 20}, num1_cols={15, 16, 17, 18},
           first=first)


def _sheet_dashboard(wb, db, params, first=False):
    d = pur.records_dashboard(
        db, params["start"], params["end"], vendor=params.get("vendor"),
        mclass=params.get("mclass"), q=params.get("q"), team=params.get("team"))
    # 요약
    summ_h = ["항목", "값"]
    summ_rows = [
        ["기간", f"{d['start']} ~ {d['end']}"],
        ["전표수", d["line_count"]],
        ["공급가 합계", d["total_supply"]],
        ["합계(VAT포함)", d["total_amount"]],
        ["거래처수", d["vendor_count"]],
        ["품목수", d["item_count"]],
        ["기간 매출", d["sales"]],
        ["매입/매출 비율(%)", d["purchase_to_sales_ratio"]],
    ]
    _sheet(wb, "대시보드요약", summ_h, summ_rows, money_cols=set(), first=first)
    # 구분별
    _sheet(wb, "구분별", ["구분", "공급가", "전표수"],
           [[r["mclass"], r["supply"], r["lines"]] for r in d["by_class"]],
           money_cols={1})
    # 거래처별
    _sheet(wb, "거래처별", ["거래처", "공급가"],
           [[r["vendor"], r["supply"]] for r in d["by_vendor"]], money_cols={1})
    # 품목별
    _sheet(wb, "품목별", ["품목코드", "품목명", "구분", "공급가", "수량"],
           [[r.get("item_code"), r.get("item_name"), r.get("mclass"), r["supply"], r["qty"]]
            for r in d["by_item"]], money_cols={3}, num1_cols={4})
    # 월별
    _sheet(wb, "월별추이", ["월", "공급가"],
           [[r["month"], r["supply"]] for r in d["by_month"]], money_cols={1})
    # 일별
    _sheet(wb, "일별추이", ["일자", "공급가"],
           [[r["date"], r["supply"]] for r in d["by_day"]], money_cols={1})


def _sheet_monthly(wb, db, params, first=False):
    year = params.get("year") or date.today().year
    by = params.get("by") or "vendor"
    d = pur.records_monthly_matrix(db, year, by=by, team=params.get("team"),
                                   mclass=params.get("mclass"))
    label_h = "거래처" if by == "vendor" else "품목명"
    months = [f"{m}월" for m in range(1, 13)]
    headers = [label_h] + (["품목코드"] if by == "material" else []) + months + ["합계"]
    rows = []
    for g in d["rows"]:
        base = [g["label"]] + ([g.get("code")] if by == "material" else [])
        rows.append(base + g["months"] + [g["total"]])
    # 월별 합계행
    lead = ["합계"] + ([""] if by == "material" else [])
    sum_row = lead + d["month_totals"] + [d["grand_total"]]
    money_start = 2 if by == "material" else 1
    money_cols = set(range(money_start, money_start + 13))  # 12개월 + 합계
    _sheet(wb, f"{year}년_월별", headers, rows, money_cols=money_cols,
           sum_row=sum_row, first=first)


def _sheet_ap(wb, db, params, first=False):
    d = pur.ap_aging(db, asof=params.get("asof"), start=params.get("start"))
    buckets = d["bucket_order"]
    headers = ["우선순위", "거래처", "정산조건", "매입액", "지급액", "잔액", "연체액",
               "평균연체일", "최장연체일", "평균지급소요일", "최초만기"] + buckets
    rows = []
    for v in d["vendors"]:
        vb = v.get("buckets", {})
        rows.append([
            v.get("priority"), v.get("vendor"), v.get("term_label") or "미설정",
            v.get("payable"), v.get("paid"), v.get("balance"), v.get("overdue"),
            v.get("avg_days_overdue"), v.get("max_days_overdue"),
            v.get("avg_pay_days"), v.get("earliest_due"),
        ] + [vb.get(b, 0) for b in buckets])
    t = d["totals"]
    sum_row = ["", "합계", "", t["payable"], t["paid"], t["balance"], t["overdue"],
               t.get("avg_days_overdue"), t.get("max_days_overdue"), t.get("avg_pay_days"), ""] + \
              [d["bucket_totals"].get(b, 0) for b in buckets]
    money_cols = set(range(3, 7)) | set(range(11, 11 + len(buckets)))
    _sheet(wb, "매입채무", headers, rows, money_cols=money_cols, sum_row=sum_row, first=first)


def _sheet_bom(wb, db, params, first=False):
    d = pur.bom_purchase_mapping(db)
    headers = ["유형", "erp코드", "품명", "공급처", "기준", "마스터가", "구매가",
               "괴리%", "거래처", "최근구매일", "매입횟수", "BOM사용", "상태"]
    flagmap = {"no_purchase": "구매없음", "stale": "괴리", "unit_check": "kg환산불가", "unused": "미사용"}
    rows = []
    for x in d["items"]:
        rows.append([
            "원재료" if x["type"] == "raw" else "부자재", x.get("erp_code"), x.get("name"),
            x.get("supplier"), x.get("basis"), x.get("master_price"), x.get("buy_price"),
            x.get("gap_pct"), x.get("vendor"), x.get("last_date"), x.get("buy_count"),
            x.get("bom_uses"), ",".join(flagmap.get(f, f) for f in x.get("flags", [])),
        ])
    _sheet(wb, "BOM매핑", headers, rows, money_cols={5, 6}, num1_cols={7}, first=first)


_BUILDERS = {
    "records": _sheet_records,
    "price": _sheet_price,
    "dashboard": _sheet_dashboard,
    "monthly": _sheet_monthly,
    "ap": _sheet_ap,
    "bom": _sheet_bom,
}
_LABELS = {
    "records": "실적조회", "price": "단가추이", "dashboard": "실적대시보드",
    "monthly": "월별매트릭스", "ap": "매입채무", "bom": "BOM매핑", "all": "구매관리전체",
}


def build_workbook(db: Session, kind: str, params: dict) -> io.BytesIO:
    wb = Workbook()
    if kind == "all":
        order = ["records", "price", "dashboard", "monthly", "ap", "bom"]
        for i, k in enumerate(order):
            try:
                _BUILDERS[k](wb, db, params, first=(i == 0))
            except Exception as e:  # 한 시트가 실패해도 나머지는 뽑는다
                ws = wb.active if i == 0 else wb.create_sheet()
                ws.title = _LABELS[k][:31]
                ws.cell(row=1, column=1, value=f"생성 실패: {str(e)[:200]}")
    else:
        builder = _BUILDERS.get(kind)
        if not builder:
            raise ValueError(f"알 수 없는 kind: {kind}")
        builder(wb, db, params, first=True)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def filename_for(kind: str) -> str:
    label = _LABELS.get(kind, kind)
    return f"구매관리_{label}_{date.today().isoformat()}.xlsx"
